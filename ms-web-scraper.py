import time
start_time = time.time()
#-----------------------------------------------------------------------------------------------------------------------------------------------------------------------#
#A script that scrapes "Trailing Returns GBP : Total Returns" and "Quarterly Returns" of Funds from website MORNINGSTAR
#To install the used libraries enter in cmd for e.g. : pip install pandas ---> If doesnt work/error occurs try uninstalling python then re-installing for all users (worked for me)
#James Howell-Smith

#[WARN] make sure all used libraries (imports) are installed
#[WARN] if "requests.get(url)" fails then webscraping using "BeautifulSoup" is not possible for the requested url
#[WARN] if 404 ERROR occurs recommend just rerunning - will be flagged in Shell and the row for respective failed fund code will not appear in MSFD.xlsx

#[NOTE] "Time" and "Time Scraped" are different e.g. Time : 1 day, 1 week, 1 month...; Time Scraped : 2021-08-09 11:48:42.9
#[NOTE] if program running whilst time changes from 00:00 to 00:01, not all duplicate columns will be dropped if same fund is called multiple times - can be fixed by removing "Time Scraped"

#[PATCH NOTES]
#[Added v1.9.1] Added AJ Bell urls for each fund in final excel file 'MSFD.xlsx' - do not need 'codesMSFD' to cotain any fund hyperlinks to work
#To comment selected lines with mouse do "alt" + 3 and to uncomment "alt" + 4 in IDLE

from bs4 import BeautifulSoup
import requests
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Fill
from datetime import datetime

frames = []
framesNo2016 = []
failed = []
failed2 = []
flagged = []
print("Running... (~25sec. for 70 codes in seperate excel workbook: codesMSFD.xlsx)")
print("")
##print("[NOTE] each element in the ID column is a hyperlink to the respective fund website via MORNINGSTAR")
##print("")
print("If a fund code is [FLAGGED] then the respective fund code will not be added to the final excel spreadsheet MSFD.xlsx")
print("")
print("********************************************************************************")

def scrape (ID):
    if ID == None:
        return
   
    
    url = f'https://www.morningstar.co.uk/uk/funds/snapshot/snapshot.aspx?id={ID}&tab=1' #Scrapes data off url by replacing {ID} with the fund code

    page = requests.get(url) #[WARN] if this fails then webscraping using "BeautifulSoup" is not possible for the requested url
    soup = BeautifulSoup(page.text, 'html.parser') #Pulls all html of website with respective url

    #print(page.url)
    
    if "SecurityToken" in page.url:
        print("")
        print("[FLAGGED] - scraping data from: " + ID + " is not possible - security token is present")
        flagged.append(ID)
        return
    
    table = soup.find('table', {'class' : 'snapshotTextColor snapshotTextFontStyle snapshotTable returnsTrailingTable'}) #Finds the table conaing data for returns
    name = soup.find('h1').text #Respective name of the fund
    time_scraped = str(datetime.now())
    time_scraped = time_scraped[:-16:]

    returns = [time_scraped, name, ID] #Return data stored here for fund
    returnsFormatted = [] #Asterisk removed from each element
    titles = ["Time Scraped", "Company", "ID"] #Time/title data stored here for fund 
    titlesFormatted = [] #Asterisk removed from each element (avoids duplicated columns with asterisk e.g. 10yrs, ..., 10yrs*)
    dict1 = {} #Where titlesFormatted and returnsFormatted linked together via dictionary method
    years = []
    times = []
    Qdata = []
    Qdata1 = []
    Qdata2 = []
    Qdata3 = []
    Qdata4 = []
    Qdata5 = []
    Qdata6 = []
    QdataSorted = []
   
    try:
        for row in table.find_all('tr')[1:]: #[WARN] if this fails, table data likely different/unusual format (or data is null when try-except block removed)
            try:
                data_time = row.find_all('td', {'class' : "col1 label"}) #Scrapes column 1 of table which is the data for time (will be appended to "titles")
            except:
                pass
            try:
                data_return = row.find_all('td', {'class' : "col2 value number"}) #Scrapes column 2 of table which is the data for percentage returns (will be appended to "returns")
            except:
                pass
            row_time = [td.text.strip() for td in data_time] #Converts from HTML to text for time
            row_time = [item for item in row_time if (item != None)] #Ignores NULL data for time
            row_return = [td.text.strip() for td in data_return] #Converts from HTML to text for return
            row_return = [item for item in row_return if (item != None)] #ignores NULL data for return

            try:
                titles.append(row_time[0]) #time appended to "titles"; data is list with length 1 hence append index [0]
            except:
                pass
            try:
                returns.append(row_return[0]) #returns appended to "returns"
            except:
                pass
    except:
        print("")
        print("[WARN] " + name + " " + str(ID) + " did not manage scrape data properly") # - will be given the option to reattempt scrape") #Triggers when scraping unsuccessful - usally 404 ERROR so attempts to scrape again
        #failed.append(ID)
        return ID
    try:
        table2 = soup.find('table', {'class' : 'snapshotTextColor snapshotTextFontStyle snapshotTable returnsQuarterlyTable'})
    
        for row in table2.find_all('tr')[1:]:
            Year = row.find_all('td', {'class' : "col1 label"})
            try:
                Data_Q1 = row.find_all('td', {'class' : "col2 value number"})
            except:
                pass
            try:
                Data_Q2 = row.find_all('td', {'class' : "col3 value number"})
            except:
                pass
            try:
                Data_Q3 = row.find_all('td', {'class' : "col4 value number"})
            except:
                pass
            try:
                Data_Q4 = row.find_all('td', {'class' : "col5 value number"})
            except:
                pass
            
            year_Q1 = [td.text.strip() for td in Year]
            year_Q1 = [item for item in year_Q1 if (item != None)]
             
            data_Q1 = [td.text.strip() for td in Data_Q1]
            data_Q1 = [item for item in data_Q1 if (item != None)]
            data_Q2 = [td.text.strip() for td in Data_Q2]
            data_Q2 = [item for item in data_Q2 if (item != None)]
            data_Q3 = [td.text.strip() for td in Data_Q3]
            data_Q3 = [item for item in data_Q3 if (item != None)]
            data_Q4 = [td.text.strip() for td in Data_Q4]
            data_Q4 = [item for item in data_Q4 if (item != None)]

            try:
                years.append(year_Q1[0] + "Q4")

            except:
                pass
            try:
                years.append(year_Q1[0] + "Q3")

            except:
                pass
            try:
                years.append(year_Q1[0] + "Q2")

            except:
                pass
            try:
                years.append(year_Q1[0] + "Q1")

            except:
                pass
            try:
                Qdata.append(data_Q1[0])
                
            except:
                pass
            try:
                Qdata.append(data_Q2[0])
                
            except:
                pass
            try:
                Qdata.append(data_Q3[0])
                
            except:
                pass
            try:
                Qdata.append(data_Q4[0])
                
            except:
                pass
            if len(failed2) > 0:
                for i in range(0, len(failed2)):
                    if failed2[i] == ID:
                        failed2.remove(ID)
                        
    except:
        pass
            
    yearsCorrected = years[::-1]
    
    try:
        for item in yearsCorrected:
            item = item.strip("*")
            titles.append(item)
    except:
        pass
    try:
        for i in range(0, 4):
            Qdata1.append(Qdata[i])
    except:
        pass
    try:
        for i in range(4, 8):
            Qdata2.append(Qdata[i])
    except:
        pass
    try:
        for i in range(8, 12):
            Qdata3.append(Qdata[i])
    except:
        pass
    try:
        for i in range(12, 16):
            Qdata4.append(Qdata[i])
    except:
        pass
    try:
        for i in range(16, 20):
            Qdata5.append(Qdata[i])
    except:
        pass
    try:
        for i in range(20, 24):
            Qdata6.append(Qdata[i])
    except:
        pass
        
    QdataSorted = Qdata6 + Qdata5 + Qdata4 + Qdata3 + Qdata2 + Qdata1

    returns = returns + QdataSorted
    
    returnsFormatted.append(time_scraped)  
    for i in range(1, len(returns)):
        returnsNew = returns[i].strip("*") #Removes asetrisk from returns
        try:
            returnsNew = float(returnsNew)
        except:
            pass
        returnsFormatted.append(returnsNew)
        
    titlesFormatted.append("Time Scraped")   
    for i in range(1, len(titles)):
        titlesNew = titles[i].strip("*") #Removes asterisk from titles - avoids duplicate columns containing asterisk in the title
        titlesFormatted.append(titlesNew)
        
    for i in range(0, len(titlesFormatted)):
        dict1[titlesFormatted[i]] = returnsFormatted[i] #"titlesFormatted" : key, "returnsFormatted" : value - now linked to one another
    data = [dict1]
    df = pd.DataFrame(data) #All data appended to a dataframe
    #df.to_excel(name + ".xlsx") #Uncomment if you want individual excel files with respective Fund data

    if '2016Q1' in df:
        frames.append(df) #Local dataframe appended to 'master' global dataframe list "frames"

    elif '2016Q1' not in df:
        framesNo2016.append(df) #Local dataframe appended to 'master' global dataframe list "frames"

    else: print("[FATAL] The data for " + ID + " has data for 2016Q1 but doesn't have data for 2016Q1 at the same time - REMOVE: " + ID + " and rerun")
        
def hyperlink ():
    
    wrkbk = openpyxl.load_workbook("MSFD.xlsx") #Name of excel file with fund codes
  
    sh = wrkbk.active #Loads up worksheet of codesMSFD
    sh.insert_cols(1)
    for i in range(2, sh.max_row+1):
        for j in range(4, 5):
            cell_obj = sh.cell(row = i, column = j)
            cell_obj2 = 'https://www.youinvest.co.uk/market-research-redirect?id=' + cell_obj.value + '&SecurityToken=' + cell_obj.value + '%5D2%5D1%5DFXALL%24%24ALL_1392&ClientFund=1&LanguageId=en-GB&CurrencyId=GBP&UniverseId=FXALL%24%24ALL_1392&BaseCurrencyId=GBP&ms-redirect-path=%2F1c6qh1t6k9%2Fsnapshot%2Fsnapshot.aspx'
            sh.cell(row = i, column = 1).value = cell_obj2
            sh.cell(row = i, column = 1).hyperlink = cell_obj2
            ##sh.cell(row = i, column = j).hyperlink = 'https://www.morningstar.co.uk/uk/funds/snapshot/snapshot.aspx?id=' + cell_obj.value + '&tab=1'
    wrkbk.save('MSFD.xlsx')
    
    
#-----------------------------------------------------------------------------------------------------------------------------------------------------------------------# End of function      

##codes = ['F0GBR0506U', #Scrape Fund data from list
##         'F00000NJPC',
##         'F00000XGIE']
##for code in codes: #Iterates through list "codes"
##    scrape(code) #Calls the function "scrape" to scrape the data for MSFD for each element, "code", in "codes"
    
##scrape('F0GBR0506U')
##scrape('F000014C09')
##scrape('F00000PAUG')


codes = []
wrkbk = openpyxl.load_workbook("codesMSFD.xlsx") #Name of excel file with fund codes
  
sh = wrkbk.active #Loads up worksheet of codesMSFD
  
for row in sh.iter_rows(min_row=1, min_col=1, max_row=sh.max_row, max_col=1): #Parameters set to all first column in codesMSFD
    for cell in row:
        codes.append(cell.value) #Appends all codes in codesMSFD to local list "codes"    
for code in codes: #Iterates through list "codes" 
    outID = scrape(code) #Calls the function "scrape" to scrape the data for MSFD for each element, "code", in "codes"
    if outID is not None:
        failed.append(outID)
        

##failed2 = failed
##for i in range(0, len(failed2)):
##    scrape(failed2[i])

##for i in range(0, len(failed2)):
##    while True:
##        try: 
##            scrape(failed2[i])
##        except:
##            continue
##        break

print("")
print("********************************************************************************")
print("")          

if len(failed) > 0:
    print("")
    print("Codes that failed: " + str(list(set(failed))))
if len(flagged) > 0:
    print("")
    print("[FLAGGED]: " + str(list(set(flagged))))
    
for fail in failed:
    again = 'y'
    while again == 'y':
        print("")
        again = input(fail + " failed to webscrape: retry again? (if no [WARN] then scrape for respective code was successful) (y/n): ")
        if again == 'y':
            outID = scrape(fail)
            if outID is not None:
                again = 'y'
            else: again = 'n'
    
if len(framesNo2016) != 0:
    frames = frames + framesNo2016
    
result = pd.concat(frames, axis=0, join='outer') #Concatenates all dataframes together in "frames"
result = result.drop_duplicates()
result.reset_index(drop = True, inplace = True)
result.to_excel("MSFD.xlsx") #Appends result to excel spreadsheet "MSFD.xlsx"

book = openpyxl.load_workbook('MSFD.xlsx') #Comment out this block if you want the dataframe index column on "MSFD.xlsx"
sheet = book['Sheet1']
sheet.delete_cols(1)
book.save('MSFD.xlsx')
hyperlink()

wb = openpyxl.load_workbook("MSFD.xlsx") #Name of excel file with fund codes
ws1 = wb.create_sheet('Flagged Codes')
for s in range(len(wb.sheetnames)):
    if wb.sheetnames[s] == 'Flagged Codes':
        break
wb.active = s
sheet = wb.active
for i in range(0,len(flagged)):
    sheet.cell(row = i + 1, column = 1).value = flagged[i]
for s in range(len(wb.sheetnames)):
    if wb.sheetnames[s] == 'Sheet1':
        break
wb.active = s
wb.active.title = "MSFD"
wb.save("MSFD.xlsx")

print("")
print("[Done] %s seconds" % (time.time() - start_time)) #Outputs time for program to complete
